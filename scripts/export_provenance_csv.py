"""Generate a provenance-tagged CSV/XLSX of the 50 SFO entities.

Every key cell (email, phone, LinkedIn, AUM, website) carries its
confidence/provenance status as a suffix:
  value [verified]
  value [catch_all]
  value [unresolved]
  — [undisclosed]
"""

import csv
import json
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / "data" / "sfo_enriched.json"
OUT_CSV = ROOT / "data" / "sfo_provenance_export.csv"
OUT_XLSX = ROOT / "data" / "sfo_provenance_export.xlsx"


def tag(value: str | None, confidence: str | None) -> str:
    """Format a cell as 'value [confidence]'."""
    if not value or value == "Unresolved":
        return "— [unresolved]"
    conf = (confidence or "unverified").lower()
    return f"{value} [{conf}]"


def flatten(entity: dict) -> dict:
    """Flatten one entity into a single CSV row with provenance tags."""
    principals = entity.get("principals", [])
    contacts = entity.get("contacts", [])
    signals = entity.get("signals", [])

    # Contact extraction
    emails = [c for c in contacts if c.get("type") == "email"]
    phones = [c for c in contacts if c.get("type") == "phone"]
    linkedins = []
    for p in principals:
        url = p.get("linkedin_url", "")
        if url and url != "Unresolved":
            linkedins.append({"value": url, "confidence": "Verified"})

    email_val = emails[0].get("value") if emails else None
    email_conf = emails[0].get("confidence") if emails else None
    phone_val = phones[0].get("value") if phones else None
    phone_conf = phones[0].get("confidence") if phones else None
    linkedin_val = linkedins[0]["value"] if linkedins else None
    linkedin_conf = linkedins[0]["confidence"] if linkedins else None

    p1 = principals[0] if principals else {}
    p2 = principals[1] if len(principals) > 1 else {}

    # Signals summary
    sig_types = list({s.get("type", "") for s in signals})
    sig_count = len(signals)

    return {
        "Entity ID": entity.get("id", ""),
        "Entity Name": entity.get("entity_name", ""),
        "Entity Type": entity.get("entity_type", ""),
        "Family Name": entity.get("family_name", ""),
        "Source of Wealth": entity.get("source_of_wealth", ""),
        "AUM (USD)": entity.get("estimated_aum_usd") or "",
        "AUM Confidence": entity.get("aum_confidence", "Undisclosed"),
        "Year Established": entity.get("year_established") or "",
        "Website": entity.get("website") or "",
        "HQ City": entity.get("hq_city", ""),
        "HQ Country": entity.get("hq_country", ""),
        "Discovery Source": entity.get("discovery_source", ""),
        "Enrichment Status": entity.get("enrichment_status", ""),
        # ── Principal 1
        "P1 Name": p1.get("full_name", ""),
        "P1 Title": p1.get("title", ""),
        "P1 LinkedIn": tag(p1.get("linkedin_url"), "Verified" if p1.get("linkedin_url") and p1["linkedin_url"] != "Unresolved" else "Unresolved"),
        # ── Principal 2
        "P2 Name": p2.get("full_name", ""),
        "P2 Title": p2.get("title", ""),
        "P2 LinkedIn": tag(p2.get("linkedin_url"), "Verified" if p2.get("linkedin_url") and p2["linkedin_url"] != "Unresolved" else "Unresolved"),
        # ── Contacts (with provenance)
        "Email": tag(email_val, email_conf),
        "Phone": tag(phone_val, phone_conf),
        "LinkedIn (Contact)": tag(linkedin_val, linkedin_conf),
        # ── Signals
        "Signal Count": sig_count,
        "Signal Types": ", ".join(sig_types) if sig_types else "",
        # ── Evidence
        "Inclusion Evidence": entity.get("inclusion_evidence", ""),
        # ── Timestamps
        "Created At": entity.get("created_at", ""),
        "Updated At": entity.get("updated_at", ""),
    }


def main():
    with open(DATA, encoding="utf-8") as f:
        raw = json.load(f)

    entities = raw.get("entities", raw) if isinstance(raw, dict) else raw

    rows = [flatten(e) for e in entities]
    fieldnames = list(rows[0].keys())

    # CSV
    with open(OUT_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
    print(f"CSV  → {OUT_CSV}  ({len(rows)} rows)")

    # XLSX (if openpyxl available)
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "SFO Provenance"

        # Header styling
        header_fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
        header_font = Font(color="f1f5f9", bold=True, size=10)
        wrap = Alignment(wrap_text=True, vertical="top")

        for col_idx, col_name in enumerate(fieldnames, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = wrap

        # Data rows
        verified_fill = PatternFill(start_color="d1fae5", end_color="d1fae5", fill_type="solid")
        catchall_fill = PatternFill(start_color="fef3c7", end_color="fef3c7", fill_type="solid")
        unresolved_fill = PatternFill(start_color="fee2e2", end_color="fee2e2", fill_type="solid")

        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, col_name in enumerate(fieldnames, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row_data[col_name])
                cell.alignment = wrap
                val = str(row_data[col_name])
                if "[verified]" in val.lower():
                    cell.fill = verified_fill
                elif "[catch_all]" in val.lower():
                    cell.fill = catchall_fill
                elif "[unresolved]" in val.lower():
                    cell.fill = unresolved_fill

        # Auto-width (approximate)
        for col_idx, col_name in enumerate(fieldnames, 1):
            max_len = max(len(col_name), max(len(str(r[col_name])) for r in rows))
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 2, 45)

        ws.auto_filter.ref = ws.dimensions
        wb.save(OUT_XLSX)
        print(f"XLSX → {OUT_XLSX}  ({len(rows)} rows)")
    except ImportError:
        print("openpyxl not installed — skipping XLSX (pip install openpyxl)")

    # Summary
    total = len(rows)
    verified = sum(1 for r in rows if "[verified]" in r["Email"].lower())
    catchall = sum(1 for r in rows if "[catch_all]" in r["Email"].lower())
    unresolved = sum(1 for r in rows if "[unresolved]" in r["Email"].lower())
    print(f"\nSummary: {total} entities | {verified} verified | {catchall} catch_all | {unresolved} unresolved")


if __name__ == "__main__":
    main()
